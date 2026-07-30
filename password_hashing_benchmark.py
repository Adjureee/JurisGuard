#!/usr/bin/env python3
"""
JurisGuard Password Hashing Benchmark
====================================

Benchmarks bcrypt, Argon2id, and PBKDF2-HMAC-SHA256 on the machine where
this script is executed. It generates synthetic passwords, measures latency,
throughput, and peak incremental resident memory, tests long-input handling,
calculates a Weighted Suitability Score (WSS), prints a console table, and
exports a formatted Excel audit workbook.

Local setup:
    python -m pip install -r requirements_password_benchmark.txt
    python password_hashing_benchmark.py

Google Colab setup:
    !pip install bcrypt argon2-cffi psutil tabulate openpyxl
    !python password_hashing_benchmark.py

The default run uses exactly 100 synthetic passwords and the required research
parameters. Use --quick only to validate that the script works; quick-mode
results are not suitable for the final capstone analysis.
"""

from __future__ import annotations

import argparse
import gc
import hashlib
import hmac
import importlib.metadata
import math
import multiprocessing as mp
import os
import platform
import queue
import random
import statistics
import string
import sys
import threading
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

try:
    import bcrypt
    import psutil
    from argon2 import PasswordHasher, Type
    from argon2.low_level import hash_secret_raw
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension
    from tabulate import tabulate
except ImportError as exc:
    missing = getattr(exc, "name", str(exc))
    raise SystemExit(
        f"Missing dependency: {missing}. Install requirements with:\n"
        "python -m pip install bcrypt argon2-cffi psutil tabulate openpyxl"
    ) from exc


# ---------------------------------------------------------------------------
# Research configuration
# ---------------------------------------------------------------------------
BCRYPT_ROUNDS = 12
ARGON2_TIME_COST = 2
ARGON2_MEMORY_COST_KIB = 65_536  # 64 MiB, as required by the study design
ARGON2_PARALLELISM = 1
PBKDF2_ITERATIONS = 600_000
SALT_BYTES = 16
DERIVED_KEY_BYTES = 32

SECURITY_SCORES = {
    "Argon2id": 10.0,
    "bcrypt": 8.0,
    "PBKDF2-HMAC-SHA256": 6.0,
}

WSS_WEIGHTS = {
    "security": 0.40,
    "latency": 0.35,
    "memory_fit": 0.25,
}

SOURCE_URLS = {
    "OWASP Password Storage Cheat Sheet": (
        "https://cheatsheetseries.owasp.org/cheatsheets/Password_Storage_Cheat_Sheet.html"
    ),
    "pyca/bcrypt documentation": "https://github.com/pyca/bcrypt",
    "argon2-cffi API": "https://argon2-cffi.readthedocs.io/en/stable/api.html",
    "Python hashlib PBKDF2": "https://docs.python.org/3/library/hashlib.html",
}


@dataclass(frozen=True)
class PasswordRecord:
    record_id: int
    password: str
    target_length_chars: int
    actual_length_chars: int
    utf8_length_bytes: int
    complexity: str


@dataclass
class BenchmarkResult:
    algorithm: str
    configuration: str
    sample_count: int
    average_latency_ms: float
    median_latency_ms: float
    p95_latency_ms: float
    stddev_latency_ms: float
    min_latency_ms: float
    max_latency_ms: float
    total_time_seconds: float
    throughput_hashes_per_second: float
    peak_memory_mb_per_hash: float
    security_score: float
    latency_score: float
    memory_fit_score: float
    wss: float
    correctness_check: str
    input_length_handling: str
    input_length_details: str
    rank: int = 0


# ---------------------------------------------------------------------------
# Synthetic password dataset
# ---------------------------------------------------------------------------
def _fit_to_length(text: str, length: int, pad_alphabet: str, rng: random.Random) -> str:
    """Return text truncated/padded to an exact character length."""
    if len(text) >= length:
        return text[:length]
    needed = length - len(text)
    return text + "".join(rng.choice(pad_alphabet) for _ in range(needed))


def _generate_password(length: int, complexity_index: int, rng: random.Random) -> tuple[str, str]:
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits
    symbols = "!@#$%^&*()-_=+[]{}:,.?"

    if complexity_index == 0:
        alphabet = lowercase + digits
        password = "".join(rng.choice(alphabet) for _ in range(length))
        label = "Lowercase + digits"
    elif complexity_index == 1:
        alphabet = lowercase + uppercase + digits
        required = [rng.choice(lowercase), rng.choice(uppercase), rng.choice(digits)]
        remaining = [rng.choice(alphabet) for _ in range(max(0, length - len(required)))]
        chars = required + remaining
        rng.shuffle(chars)
        password = "".join(chars[:length])
        label = "Mixed case + digits"
    elif complexity_index == 2:
        alphabet = lowercase + uppercase + digits + symbols
        required = [
            rng.choice(lowercase),
            rng.choice(uppercase),
            rng.choice(digits),
            rng.choice(symbols),
        ]
        remaining = [rng.choice(alphabet) for _ in range(max(0, length - len(required)))]
        chars = required + remaining
        rng.shuffle(chars)
        password = "".join(chars[:length])
        label = "Mixed case + digits + symbols"
    else:
        words = [
            "legal",
            "archive",
            "client",
            "secure",
            "justice",
            "record",
            "verify",
            "access",
            "audit",
            "case",
        ]
        pieces: list[str] = []
        while len("-".join(pieces)) < length:
            word = rng.choice(words)
            if rng.random() < 0.5:
                word = word.capitalize()
            pieces.append(word)
        phrase = "-".join(pieces) + str(rng.randrange(10)) + rng.choice(symbols)
        password = _fit_to_length(phrase, length, lowercase + uppercase + digits + symbols, rng)
        label = "Passphrase-style + digits + symbols"

    return password, label


def generate_dummy_passwords(count: int = 100, seed: int = 2026) -> list[PasswordRecord]:
    """
    Generate synthetic passwords spanning 8, 16, 32, and 64 characters.

    With the default count of 100, exactly 25 passwords are generated for each
    target length. The dataset is deterministic for audit reproducibility.
    """
    if count < 4:
        raise ValueError("count must be at least 4")

    rng = random.Random(seed)
    lengths = [8, 16, 32, 64]
    records: list[PasswordRecord] = []

    # Cycling lengths ensures balanced groups for 100 and near-balanced groups
    # for any smaller smoke-test count.
    for index in range(count):
        length = lengths[index % len(lengths)]
        password, complexity = _generate_password(length, index % 4, rng)
        records.append(
            PasswordRecord(
                record_id=index + 1,
                password=password,
                target_length_chars=length,
                actual_length_chars=len(password),
                utf8_length_bytes=len(password.encode("utf-8")),
                complexity=complexity,
            )
        )

    rng.shuffle(records)
    # Restore stable audit IDs after shuffling the benchmark order.
    return [
        PasswordRecord(
            record_id=i + 1,
            password=record.password,
            target_length_chars=record.target_length_chars,
            actual_length_chars=record.actual_length_chars,
            utf8_length_bytes=record.utf8_length_bytes,
            complexity=record.complexity,
        )
        for i, record in enumerate(records)
    ]


# ---------------------------------------------------------------------------
# Hash implementations
# ---------------------------------------------------------------------------
def _argon2_hasher() -> PasswordHasher:
    return PasswordHasher(
        time_cost=ARGON2_TIME_COST,
        memory_cost=ARGON2_MEMORY_COST_KIB,
        parallelism=ARGON2_PARALLELISM,
        hash_len=DERIVED_KEY_BYTES,
        salt_len=SALT_BYTES,
        type=Type.ID,
    )


_ARGON2_HASHER: PasswordHasher | None = None


def _get_argon2_hasher() -> PasswordHasher:
    global _ARGON2_HASHER
    if _ARGON2_HASHER is None:
        _ARGON2_HASHER = _argon2_hasher()
    return _ARGON2_HASHER


def hash_bcrypt(password: str) -> bytes:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=BCRYPT_ROUNDS))


def hash_argon2id(password: str) -> str:
    return _get_argon2_hasher().hash(password)


def hash_pbkdf2(password: str) -> bytes:
    return hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        os.urandom(SALT_BYTES),
        PBKDF2_ITERATIONS,
        dklen=DERIVED_KEY_BYTES,
    )


HASH_FUNCTIONS: dict[str, Callable[[str], Any]] = {
    "bcrypt": hash_bcrypt,
    "Argon2id": hash_argon2id,
    "PBKDF2-HMAC-SHA256": hash_pbkdf2,
}

CONFIGURATIONS = {
    "bcrypt": f"cost/work factor={BCRYPT_ROUNDS}",
    "Argon2id": (
        f"time_cost={ARGON2_TIME_COST}, memory_cost={ARGON2_MEMORY_COST_KIB} KiB, "
        f"parallelism={ARGON2_PARALLELISM}"
    ),
    "PBKDF2-HMAC-SHA256": f"iterations={PBKDF2_ITERATIONS:,}, salt={SALT_BYTES} bytes",
}


# ---------------------------------------------------------------------------
# Validation and benchmark helpers
# ---------------------------------------------------------------------------
def percentile(values: Iterable[float], percentile_value: float) -> float:
    ordered = sorted(values)
    if not ordered:
        return 0.0
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * percentile_value
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] * (1 - fraction) + ordered[upper] * fraction


def correctness_check(algorithm: str, password: str) -> str:
    """Perform a single hash/verification test outside the timed benchmark."""
    if algorithm == "bcrypt":
        encoded = password.encode("utf-8")
        digest = bcrypt.hashpw(encoded, bcrypt.gensalt(rounds=BCRYPT_ROUNDS))
        return "PASS" if bcrypt.checkpw(encoded, digest) else "FAIL"

    if algorithm == "Argon2id":
        hasher = _argon2_hasher()
        digest = hasher.hash(password)
        return "PASS" if hasher.verify(digest, password) else "FAIL"

    if algorithm == "PBKDF2-HMAC-SHA256":
        salt = os.urandom(SALT_BYTES)
        encoded = password.encode("utf-8")
        digest = hashlib.pbkdf2_hmac(
            "sha256", encoded, salt, PBKDF2_ITERATIONS, dklen=DERIVED_KEY_BYTES
        )
        candidate = hashlib.pbkdf2_hmac(
            "sha256", encoded, salt, PBKDF2_ITERATIONS, dklen=DERIVED_KEY_BYTES
        )
        return "PASS" if hmac.compare_digest(digest, candidate) else "FAIL"

    raise KeyError(f"Unknown algorithm: {algorithm}")


def benchmark_latencies(
    algorithm: str, passwords: list[PasswordRecord]
) -> tuple[list[float], float]:
    hash_function = HASH_FUNCTIONS[algorithm]

    # One warm-up hash reduces one-time import and allocator effects.
    hash_function(passwords[0].password)
    gc.collect()

    latencies_ms: list[float] = []
    total_start = time.perf_counter_ns()
    for record in passwords:
        start = time.perf_counter_ns()
        hash_function(record.password)
        end = time.perf_counter_ns()
        latencies_ms.append((end - start) / 1_000_000)
    total_seconds = (time.perf_counter_ns() - total_start) / 1_000_000_000
    return latencies_ms, total_seconds


class _RSSMonitor:
    """Sample process resident memory, including native library allocations."""

    def __init__(self, interval_seconds: float = 0.0005) -> None:
        self.interval_seconds = interval_seconds
        self._stop_event = threading.Event()
        self._thread: threading.Thread | None = None
        self._process = psutil.Process(os.getpid())
        self.baseline_rss = 0
        self.peak_rss = 0

    def _sample(self) -> None:
        while not self._stop_event.is_set():
            try:
                rss = self._process.memory_info().rss
                if rss > self.peak_rss:
                    self.peak_rss = rss
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                break
            time.sleep(self.interval_seconds)

    def start(self) -> None:
        self._thread = threading.Thread(target=self._sample, daemon=True)
        self._thread.start()
        time.sleep(0.01)
        self.baseline_rss = self._process.memory_info().rss
        self.peak_rss = self.baseline_rss

    def stop(self) -> float:
        self._stop_event.set()
        if self._thread is not None:
            self._thread.join(timeout=2)
        try:
            self.peak_rss = max(self.peak_rss, self._process.memory_info().rss)
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass
        return max(0.0, (self.peak_rss - self.baseline_rss) / (1024 * 1024))


def _memory_worker(algorithm: str, password: str, result_queue: mp.Queue) -> None:
    """Measure one hash in a fresh process so allocator reuse cannot hide peaks."""
    try:
        # Construct reusable objects before taking the baseline.
        hasher = _argon2_hasher() if algorithm == "Argon2id" else None
        gc.collect()
        monitor = _RSSMonitor()
        monitor.start()
        start = time.perf_counter_ns()

        if algorithm == "bcrypt":
            bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=BCRYPT_ROUNDS))
        elif algorithm == "Argon2id":
            assert hasher is not None
            hasher.hash(password)
        elif algorithm == "PBKDF2-HMAC-SHA256":
            hashlib.pbkdf2_hmac(
                "sha256",
                password.encode("utf-8"),
                os.urandom(SALT_BYTES),
                PBKDF2_ITERATIONS,
                dklen=DERIVED_KEY_BYTES,
            )
        else:
            raise KeyError(f"Unknown algorithm: {algorithm}")

        elapsed_ms = (time.perf_counter_ns() - start) / 1_000_000
        peak_delta_mb = monitor.stop()
        result_queue.put({"ok": True, "peak_delta_mb": peak_delta_mb, "elapsed_ms": elapsed_ms})
    except Exception as exc:  # pragma: no cover - defensive subprocess reporting
        result_queue.put({"ok": False, "error": repr(exc)})


def measure_peak_memory_mb(
    algorithm: str, password: str, repetitions: int = 3, timeout_seconds: int = 120
) -> tuple[float, list[float]]:
    context = mp.get_context("spawn")
    measurements: list[float] = []

    for _ in range(repetitions):
        result_queue: mp.Queue = context.Queue()
        process = context.Process(target=_memory_worker, args=(algorithm, password, result_queue))
        process.start()
        process.join(timeout=timeout_seconds)

        if process.is_alive():
            process.terminate()
            process.join(timeout=5)
            raise TimeoutError(f"Memory benchmark timed out for {algorithm}")

        try:
            payload = result_queue.get(timeout=5)
        except queue.Empty as exc:
            raise RuntimeError(f"Memory worker returned no result for {algorithm}") from exc

        if not payload.get("ok"):
            raise RuntimeError(f"Memory worker failed for {algorithm}: {payload.get('error')}")
        measurements.append(float(payload["peak_delta_mb"]))

    return max(measurements), measurements


def test_input_length_handling() -> dict[str, dict[str, str]]:
    """Test two inputs that share the first 72 bytes but differ afterward."""
    prefix_72 = b"A" * 72
    long_a = prefix_72 + b"-FIRST-SUFFIX"
    long_b = prefix_72 + b"-SECOND-SUFFIX"
    exact_72 = prefix_72
    results: dict[str, dict[str, str]] = {}

    # bcrypt: current pyca/bcrypt raises ValueError; older versions may silently truncate.
    bcrypt_salt = bcrypt.gensalt(rounds=BCRYPT_ROUNDS)
    exact_accepted = False
    try:
        bcrypt.hashpw(exact_72, bcrypt_salt)
        exact_accepted = True
    except Exception:
        exact_accepted = False

    try:
        digest_a = bcrypt.hashpw(long_a, bcrypt_salt)
        digest_b = bcrypt.hashpw(long_b, bcrypt_salt)
        if hmac.compare_digest(digest_a, digest_b):
            handling = "FAIL: silently truncates/collides after 72 bytes"
            details = (
                "Two distinct passwords sharing the first 72 bytes produced the same hash "
                "with one fixed salt. Enforce a <=72-byte policy or use a carefully designed "
                "pre-hash scheme."
            )
        else:
            handling = "PASS: accepts >72 bytes and distinguishes suffixes"
            details = "The installed implementation distinguished the two long inputs."
    except ValueError as exc:
        handling = "LIMIT ENFORCED: rejects inputs >72 bytes"
        details = (
            f"The installed bcrypt library raised ValueError ({exc}). This avoids silent "
            "truncation but requires application-level length validation."
        )
    except Exception as exc:
        handling = f"ERROR: {type(exc).__name__}"
        details = repr(exc)

    results["bcrypt"] = {
        "exact_72": "Accepted" if exact_accepted else "Rejected",
        "long_input_bytes": str(len(long_a)),
        "handling": handling,
        "details": details,
    }

    # Argon2id low-level test with a fixed salt to prove the suffix affects the result.
    fixed_salt = b"JurisGuardSalt01"  # 16 bytes, synthetic benchmark salt
    argon_a = hash_secret_raw(
        secret=long_a,
        salt=fixed_salt,
        time_cost=ARGON2_TIME_COST,
        memory_cost=ARGON2_MEMORY_COST_KIB,
        parallelism=ARGON2_PARALLELISM,
        hash_len=DERIVED_KEY_BYTES,
        type=Type.ID,
    )
    argon_b = hash_secret_raw(
        secret=long_b,
        salt=fixed_salt,
        time_cost=ARGON2_TIME_COST,
        memory_cost=ARGON2_MEMORY_COST_KIB,
        parallelism=ARGON2_PARALLELISM,
        hash_len=DERIVED_KEY_BYTES,
        type=Type.ID,
    )
    argon_distinct = not hmac.compare_digest(argon_a, argon_b)
    results["Argon2id"] = {
        "exact_72": "Accepted",
        "long_input_bytes": str(len(long_a)),
        "handling": "PASS: full long input affects hash" if argon_distinct else "FAIL: collision",
        "details": (
            "Two inputs sharing the first 72 bytes produced different Argon2id outputs "
            "with the same salt."
        ),
    }

    pbkdf2_a = hashlib.pbkdf2_hmac(
        "sha256", long_a, fixed_salt, PBKDF2_ITERATIONS, dklen=DERIVED_KEY_BYTES
    )
    pbkdf2_b = hashlib.pbkdf2_hmac(
        "sha256", long_b, fixed_salt, PBKDF2_ITERATIONS, dklen=DERIVED_KEY_BYTES
    )
    pbkdf2_distinct = not hmac.compare_digest(pbkdf2_a, pbkdf2_b)
    results["PBKDF2-HMAC-SHA256"] = {
        "exact_72": "Accepted",
        "long_input_bytes": str(len(long_a)),
        "handling": "PASS: full long input affects hash" if pbkdf2_distinct else "FAIL: collision",
        "details": (
            "Two inputs sharing the first 72 bytes produced different PBKDF2 outputs "
            "with the same salt."
        ),
    }

    return results


# ---------------------------------------------------------------------------
# Scoring model
# ---------------------------------------------------------------------------
def latency_score(average_latency_ms: float, target_latency_ms: float) -> float:
    """
    Exponential closeness score in [0, 10], maximized at the target latency.

    Score = 10 * exp(-|measured - target| / target)

    This is symmetric around the target and avoids an arbitrary hard cutoff.
    """
    if target_latency_ms <= 0:
        raise ValueError("target_latency_ms must be greater than zero")
    return 10.0 * math.exp(-abs(average_latency_ms - target_latency_ms) / target_latency_ms)


def memory_fit_score(peak_memory_mb: float, memory_budget_mb: float) -> float:
    """
    Award 10 points while the measured incremental peak fits the configured
    per-hash memory budget; decrease proportionally when it exceeds the budget.
    """
    if memory_budget_mb <= 0:
        raise ValueError("memory_budget_mb must be greater than zero")
    if peak_memory_mb <= memory_budget_mb:
        return 10.0
    return max(0.0, min(10.0, 10.0 * memory_budget_mb / peak_memory_mb))


def calculate_wss(security: float, latency: float, memory_fit: float) -> float:
    return (
        WSS_WEIGHTS["security"] * security
        + WSS_WEIGHTS["latency"] * latency
        + WSS_WEIGHTS["memory_fit"] * memory_fit
    )


# ---------------------------------------------------------------------------
# Environment metadata
# ---------------------------------------------------------------------------
def package_version(distribution_name: str) -> str:
    try:
        return importlib.metadata.version(distribution_name)
    except importlib.metadata.PackageNotFoundError:
        return "Unknown"


def environment_metadata() -> dict[str, Any]:
    virtual_memory = psutil.virtual_memory()
    cpu_name = platform.processor() or platform.machine() or "Unknown"
    return {
        "Run timestamp (UTC)": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "Host name": platform.node() or "Unknown",
        "Operating system": platform.platform(),
        "Python version": platform.python_version(),
        "Python implementation": platform.python_implementation(),
        "CPU": cpu_name,
        "Physical CPU cores": psutil.cpu_count(logical=False) or "Unknown",
        "Logical CPU cores": psutil.cpu_count(logical=True) or "Unknown",
        "Total system RAM (GiB)": round(virtual_memory.total / (1024**3), 2),
        "bcrypt version": package_version("bcrypt"),
        "argon2-cffi version": package_version("argon2-cffi"),
        "psutil version": package_version("psutil"),
        "openpyxl version": package_version("openpyxl"),
        "tabulate version": package_version("tabulate"),
    }


# ---------------------------------------------------------------------------
# Excel audit workbook
# ---------------------------------------------------------------------------
DARK_BLUE = "17365D"
MEDIUM_BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_ORANGE = "FCE4D6"
LIGHT_RED = "F4CCCC"
LIGHT_PURPLE = "E4DFEC"
LIGHT_TEAL = "DDEBF7"
LIGHT_GRAY = "E7E6E6"
WHITE = "FFFFFF"
BLACK = "000000"
GREEN_TEXT = "008000"
BLUE_TEXT = "0000FF"
GRAY_TEXT = "666666"
PURPLE_TEXT = "7030A0"
ORANGE_TEXT = "C65911"


def _style_title(ws, title: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _style_header_row(ws, row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = ws.cell(row, column)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=MEDIUM_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32


def _set_column_widths(ws, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width


def _thin_top_border() -> Border:
    return Border(top=Side(style="thin", color="808080"))


def build_recommendation(results: list[BenchmarkResult], target_latency_ms: float, memory_budget_mb: float) -> str:
    winner = min(results, key=lambda item: item.rank)
    caveat = ""
    if winner.algorithm == "bcrypt":
        caveat = (
            " Because bcrypt has a 72-byte input boundary, deployment must enforce an "
            "appropriate byte-length policy or use a formally reviewed pre-hashing design."
        )
    elif winner.algorithm == "PBKDF2-HMAC-SHA256":
        caveat = (
            " PBKDF2 is especially relevant where FIPS-oriented compatibility is required, "
            "but it is not memory-hard."
        )
    else:
        caveat = (
            " Argon2id is memory-hard and has no bcrypt-style 72-byte truncation boundary in "
            "this test."
        )

    return (
        f"{winner.algorithm} ranked first with WSS={winner.wss:.3f}/10. "
        f"Its weighted components were 0.40×{winner.security_score:.3f} security + "
        f"0.35×{winner.latency_score:.3f} latency-fit + "
        f"0.25×{winner.memory_fit_score:.3f} memory-fit. The latency target was "
        f"{target_latency_ms:.1f} ms and the per-hash memory budget was "
        f"{memory_budget_mb:.1f} MB.{caveat} The recommendation is valid for the measured "
        "hardware and should be re-benchmarked on the production authentication server."
    )


def export_excel_report(
    output_path: Path,
    results: list[BenchmarkResult],
    raw_latencies: dict[str, list[float]],
    passwords: list[PasswordRecord],
    length_tests: dict[str, dict[str, str]],
    memory_replications: dict[str, list[float]],
    target_latency_ms: float,
    memory_budget_mb: float,
    seed: int,
    environment: dict[str, Any],
    quick_mode: bool,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    # Executive Summary -----------------------------------------------------
    summary = workbook.create_sheet("Executive Summary")
    summary.sheet_view.showGridLines = False
    _style_title(summary, "Password Hashing Benchmark — Master Audit Report", 8)

    summary_rows = [
        ("Study", "JurisGuard Objective 4 — Password Hashing Algorithm Selection"),
        ("Run mode", "QUICK SMOKE TEST — NOT FOR FINAL ANALYSIS" if quick_mode else "FULL RESEARCH BENCHMARK"),
        ("Dataset", f"{len(passwords)} deterministic synthetic passwords; lengths 8, 16, 32, and 64 characters"),
        ("Argon2id profile note", "Study-specified 64 MiB, t=2, p=1 profile; current OWASP minimum alternatives include 19 MiB, t=2, p=1"),
        ("Random seed", seed),
        ("WSS formula", "WSS = (0.40 × Security) + (0.35 × Latency Fit) + (0.25 × Memory Fit)"),
        ("Latency scoring", "10 × EXP(-ABS(measured_ms - target_ms) / target_ms)"),
        ("Target latency (ms)", target_latency_ms),
        ("Memory budget per hash (MB)", memory_budget_mb),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, value)
        summary.cell(row_index, 1).font = Font(bold=True, color=GRAY_TEXT)
        summary.cell(row_index, 1).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        summary.cell(row_index, 2).font = Font(color=BLUE_TEXT if row_index in (7, 10, 11) else BLACK)
        summary.cell(row_index, 2).alignment = Alignment(wrap_text=True)

    winner = min(results, key=lambda item: item.rank)
    summary["A13"] = "Winning recommendation"
    summary["A13"].font = Font(bold=True, color=WHITE)
    summary["A13"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    summary["B13"] = winner.algorithm
    summary["B13"].font = Font(size=14, bold=True, color=GREEN_TEXT)
    summary["A14"] = "Winning WSS"
    summary["A14"].font = Font(bold=True)
    summary["B14"] = winner.wss
    summary["B14"].number_format = "0.000"
    summary["A16"] = "Mathematical justification"
    summary["A16"].font = Font(bold=True, color=WHITE)
    summary["A16"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    summary.merge_cells("B16:H19")
    summary["B16"] = build_recommendation(results, target_latency_ms, memory_budget_mb)
    summary["B16"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["B16"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)

    # Small score table for chart
    score_start = 22
    summary.cell(score_start, 1, "Algorithm")
    summary.cell(score_start, 2, "WSS")
    _style_header_row(summary, score_start, 1, 2)
    for offset, result in enumerate(sorted(results, key=lambda item: item.rank), start=1):
        summary.cell(score_start + offset, 1, result.algorithm)
        summary.cell(score_start + offset, 2, result.wss)
        summary.cell(score_start + offset, 2).number_format = "0.000"

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Weighted Suitability Score"
    chart.y_axis.title = "Algorithm"
    chart.x_axis.title = "WSS (0–10)"
    chart.height = 7
    chart.width = 13
    chart.legend = None
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    data = Reference(summary, min_col=2, min_row=score_start, max_row=score_start + len(results))
    categories = Reference(summary, min_col=1, min_row=score_start + 1, max_row=score_start + len(results))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    summary.add_chart(chart, "D22")

    _set_column_widths(summary, {1: 30, 2: 32, 3: 3, 4: 15, 5: 15, 6: 15, 7: 15, 8: 15})

    # Comparison ------------------------------------------------------------
    comparison = workbook.create_sheet("Comparison")
    comparison.sheet_view.showGridLines = False
    _style_title(comparison, "Algorithm Comparison and Weighted Suitability Score", 17)
    headers = [
        "Algorithm",
        "Configuration",
        "Samples",
        "Avg Latency (ms)",
        "Median (ms)",
        "P95 (ms)",
        "Std Dev (ms)",
        "Peak Incremental Memory / Hash (MB)",
        "Throughput (hashes/s)",
        "Security Score",
        "Latency Score",
        "Memory Fit Score",
        "WSS",
        "Rank",
        "Long-Input Handling",
        "Correctness",
        "Audit Note",
    ]
    for col_index, header in enumerate(headers, start=1):
        comparison.cell(3, col_index, header)
    _style_header_row(comparison, 3, 1, len(headers))

    sorted_results = sorted(results, key=lambda item: item.rank)
    for row_index, result in enumerate(sorted_results, start=4):
        values = [
            result.algorithm,
            result.configuration,
            result.sample_count,
            result.average_latency_ms,
            result.median_latency_ms,
            result.p95_latency_ms,
            result.stddev_latency_ms,
            result.peak_memory_mb_per_hash,
            result.throughput_hashes_per_second,
            result.security_score,
            None,
            None,
            None,
            None,
            result.input_length_handling,
            result.correctness_check,
            result.input_length_details,
        ]
        for col_index, value in enumerate(values, start=1):
            comparison.cell(row_index, col_index, value)

        # Transparent Excel formulas mirror the Python scoring implementation.
        comparison.cell(
            row_index, 11,
            f"=10*EXP(-ABS(D{row_index}-'Executive Summary'!$B$10)/'Executive Summary'!$B$10)",
        )
        comparison.cell(
            row_index, 12,
            f"=IF(H{row_index}<='Executive Summary'!$B$11,10,10*'Executive Summary'!$B$11/H{row_index})",
        )
        comparison.cell(row_index, 13, f"=0.40*J{row_index}+0.35*K{row_index}+0.25*L{row_index}")
        comparison.cell(row_index, 14, f"=RANK.EQ(M{row_index},$M$4:$M$6,0)")

        # Measured/imported values in green; static security score in gray;
        # logic-derived scores in purple.
        for col_index in range(3, 10):
            comparison.cell(row_index, col_index).font = Font(color=GREEN_TEXT)
        comparison.cell(row_index, 10).font = Font(color=GRAY_TEXT)
        for col_index in range(11, 15):
            comparison.cell(row_index, col_index).font = Font(color=PURPLE_TEXT)
        if "FAIL" in result.input_length_handling or "LIMIT" in result.input_length_handling:
            comparison.cell(row_index, 15).fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
        if result.correctness_check != "PASS":
            comparison.cell(row_index, 16).fill = PatternFill("solid", fgColor=LIGHT_RED)
        if result.rank == 1:
            for col_index in range(1, len(headers) + 1):
                comparison.cell(row_index, col_index).fill = PatternFill("solid", fgColor=LIGHT_GREEN)

    for row in range(4, 4 + len(results)):
        for col in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
            comparison.cell(row, col).number_format = "0.000"
        comparison.cell(row, 17).alignment = Alignment(wrap_text=True, vertical="top")

    comparison.freeze_panes = "A4"
    comparison.auto_filter.ref = f"A3:Q{3 + len(results)}"
    _set_column_widths(
        comparison,
        {
            1: 24,
            2: 48,
            3: 10,
            4: 16,
            5: 14,
            6: 12,
            7: 14,
            8: 23,
            9: 18,
            10: 14,
            11: 14,
            12: 16,
            13: 10,
            14: 8,
            15: 37,
            16: 12,
            17: 70,
        },
    )

    # Raw Measurements ------------------------------------------------------
    raw = workbook.create_sheet("Raw Measurements")
    raw.sheet_view.showGridLines = False
    _style_title(raw, "Per-Password Latency Measurements", 8)
    raw_headers = [
        "Algorithm",
        "Benchmark Order",
        "Password ID",
        "Length (chars)",
        "Length (bytes)",
        "Complexity",
        "Latency (ms)",
        "Synthetic Password",
    ]
    for col_index, header in enumerate(raw_headers, start=1):
        raw.cell(3, col_index, header)
    _style_header_row(raw, 3, 1, len(raw_headers))

    row_index = 4
    password_lookup = {record.record_id: record for record in passwords}
    for algorithm in ["Argon2id", "bcrypt", "PBKDF2-HMAC-SHA256"]:
        for order, (record, latency) in enumerate(zip(passwords, raw_latencies[algorithm]), start=1):
            values = [
                algorithm,
                order,
                record.record_id,
                record.actual_length_chars,
                record.utf8_length_bytes,
                record.complexity,
                latency,
                record.password,
            ]
            for col_index, value in enumerate(values, start=1):
                raw.cell(row_index, col_index, value)
            raw.cell(row_index, 7).number_format = "0.000"
            raw.cell(row_index, 7).font = Font(color=GREEN_TEXT)
            raw.cell(row_index, 8).font = Font(color=GRAY_TEXT)
            row_index += 1

    raw.freeze_panes = "A4"
    raw.auto_filter.ref = f"A3:H{row_index - 1}"
    _set_column_widths(raw, {1: 24, 2: 15, 3: 12, 4: 14, 5: 14, 6: 34, 7: 16, 8: 68})

    # Memory Replications ---------------------------------------------------
    memory_sheet = workbook.create_sheet("Memory Measurements")
    memory_sheet.sheet_view.showGridLines = False
    _style_title(memory_sheet, "Fresh-Process Peak Incremental RSS Measurements", 5)
    memory_headers = ["Algorithm", "Replication", "Peak Delta (MB)", "Reported Peak (MB)", "Method"]
    for col_index, header in enumerate(memory_headers, start=1):
        memory_sheet.cell(3, col_index, header)
    _style_header_row(memory_sheet, 3, 1, len(memory_headers))
    row_index = 4
    result_by_algorithm = {result.algorithm: result for result in results}
    for algorithm, replications in memory_replications.items():
        for replication_index, measurement in enumerate(replications, start=1):
            memory_sheet.cell(row_index, 1, algorithm)
            memory_sheet.cell(row_index, 2, replication_index)
            memory_sheet.cell(row_index, 3, measurement)
            memory_sheet.cell(row_index, 4, result_by_algorithm[algorithm].peak_memory_mb_per_hash)
            memory_sheet.cell(
                row_index,
                5,
                "Fresh spawned process; 0.5 ms RSS sampling; delta above post-initialization baseline",
            )
            memory_sheet.cell(row_index, 3).number_format = "0.000"
            memory_sheet.cell(row_index, 4).number_format = "0.000"
            row_index += 1
    memory_sheet.freeze_panes = "A4"
    _set_column_widths(memory_sheet, {1: 24, 2: 12, 3: 18, 4: 20, 5: 78})

    # Input Handling Tests --------------------------------------------------
    input_sheet = workbook.create_sheet("Input Length Tests")
    input_sheet.sheet_view.showGridLines = False
    _style_title(input_sheet, "Truncation and Input-Length Handling Test", 6)
    input_headers = [
        "Algorithm",
        "Exactly 72 Bytes",
        "Long Test Input (bytes)",
        "Result",
        "Details",
        "Interpretation",
    ]
    for col_index, header in enumerate(input_headers, start=1):
        input_sheet.cell(3, col_index, header)
    _style_header_row(input_sheet, 3, 1, len(input_headers))
    for row_index, algorithm in enumerate(["Argon2id", "bcrypt", "PBKDF2-HMAC-SHA256"], start=4):
        test = length_tests[algorithm]
        interpretation = (
            "Application validation required"
            if algorithm == "bcrypt"
            else "No 72-byte truncation detected"
        )
        values = [
            algorithm,
            test["exact_72"],
            int(test["long_input_bytes"]),
            test["handling"],
            test["details"],
            interpretation,
        ]
        for col_index, value in enumerate(values, start=1):
            input_sheet.cell(row_index, col_index, value)
        input_sheet.cell(row_index, 4).fill = PatternFill(
            "solid",
            fgColor=LIGHT_ORANGE if algorithm == "bcrypt" else LIGHT_GREEN,
        )
        input_sheet.cell(row_index, 5).alignment = Alignment(wrap_text=True, vertical="top")
    _set_column_widths(input_sheet, {1: 24, 2: 18, 3: 22, 4: 45, 5: 85, 6: 34})

    # Dataset ---------------------------------------------------------------
    dataset_sheet = workbook.create_sheet("Synthetic Dataset")
    dataset_sheet.sheet_view.showGridLines = False
    _style_title(dataset_sheet, "Synthetic Password Dataset (Dummy Data Only)", 7)
    dataset_headers = [
        "Password ID",
        "Target Length",
        "Actual Length",
        "UTF-8 Bytes",
        "Complexity",
        "Synthetic Password",
        "Data Classification",
    ]
    for col_index, header in enumerate(dataset_headers, start=1):
        dataset_sheet.cell(3, col_index, header)
    _style_header_row(dataset_sheet, 3, 1, len(dataset_headers))
    for row_index, record in enumerate(passwords, start=4):
        values = [
            record.record_id,
            record.target_length_chars,
            record.actual_length_chars,
            record.utf8_length_bytes,
            record.complexity,
            record.password,
            "Synthetic / non-production",
        ]
        for col_index, value in enumerate(values, start=1):
            dataset_sheet.cell(row_index, col_index, value)
        dataset_sheet.cell(row_index, 6).font = Font(color=GRAY_TEXT)
    dataset_sheet.freeze_panes = "A4"
    dataset_sheet.auto_filter.ref = f"A3:G{3 + len(passwords)}"
    _set_column_widths(dataset_sheet, {1: 12, 2: 15, 3: 15, 4: 13, 5: 36, 6: 70, 7: 25})

    # Methodology -----------------------------------------------------------
    method = workbook.create_sheet("Methodology")
    method.sheet_view.showGridLines = False
    _style_title(method, "Benchmark Methodology, Scoring Model, and Sources", 6)
    methodology_rows = [
        ("Metric", "Method / Formula", "Purpose", "Audit classification"),
        (
            "Argon2id parameter profile",
            "Study requirement: m=65,536 KiB, t=2, p=1; OWASP currently publishes multiple minimum profiles including m=19,456 KiB, t=2, p=1",
            "Preserves the capstone test configuration while documenting the current guidance distinction",
            "Static/research control",
        ),
        (
            "Average latency",
            "Arithmetic mean of perf_counter_ns elapsed time for each salted hash; warm-up excluded",
            "Measures normal single-thread login hashing delay",
            "Measured",
        ),
        (
            "Throughput",
            "Number of hashes / total sequential benchmark seconds",
            "Single-thread hashes processed per second",
            "Derived",
        ),
        (
            "Peak memory per hash",
            "Maximum incremental RSS across fresh-process replications, sampled every 0.5 ms",
            "Includes native allocations such as Argon2 memory",
            "Measured/derived",
        ),
        (
            "Latency score",
            "10 × EXP(-ABS(average_latency_ms - 350) / 350), with target configurable",
            "Maximum score occurs exactly at the policy target",
            "Derived",
        ),
        (
            "Memory fit score",
            "10 when peak <= configured budget; otherwise 10 × budget / peak",
            "Tests deployment fit without rewarding excessive memory use",
            "Derived",
        ),
        (
            "WSS",
            "0.40 × Security + 0.35 × Latency Score + 0.25 × Memory Fit Score",
            "Ranks algorithms on the stated capstone decision model",
            "Derived",
        ),
        (
            "Long-input test",
            "Hashes two inputs sharing the first 72 bytes but differing afterward using the same salt",
            "Detects bcrypt truncation, rejection, or full-input handling",
            "Measured",
        ),
    ]
    for row_index, row_values in enumerate(methodology_rows, start=3):
        for col_index, value in enumerate(row_values, start=1):
            method.cell(row_index, col_index, value)
            method.cell(row_index, col_index).alignment = Alignment(wrap_text=True, vertical="top")
    _style_header_row(method, 3, 1, 4)

    source_start = 14
    method.cell(source_start, 1, "Source")
    method.cell(source_start, 2, "URL")
    method.cell(source_start, 3, "Use in benchmark")
    _style_header_row(method, source_start, 1, 3)
    for offset, (source_name, url) in enumerate(SOURCE_URLS.items(), start=1):
        method.cell(source_start + offset, 1, source_name)
        method.cell(source_start + offset, 2, url)
        method.cell(source_start + offset, 2).hyperlink = url
        method.cell(source_start + offset, 2).style = "Hyperlink"
        use_text = {
            "OWASP Password Storage Cheat Sheet": "Algorithm parameters and deployment guidance",
            "pyca/bcrypt documentation": "bcrypt API and 72-byte behavior",
            "argon2-cffi API": "Argon2id parameter semantics",
            "Python hashlib PBKDF2": "PBKDF2-HMAC-SHA256 API and salt guidance",
        }[source_name]
        method.cell(source_start + offset, 3, use_text)
    _set_column_widths(method, {1: 34, 2: 90, 3: 55, 4: 22, 5: 3, 6: 3})

    # Environment -----------------------------------------------------------
    environment_sheet = workbook.create_sheet("Environment")
    environment_sheet.sheet_view.showGridLines = False
    _style_title(environment_sheet, "Hardware and Software Environment", 4)
    environment_sheet["A3"] = "Property"
    environment_sheet["B3"] = "Value"
    environment_sheet["C3"] = "Audit Role"
    _style_header_row(environment_sheet, 3, 1, 3)
    for row_index, (key, value) in enumerate(environment.items(), start=4):
        environment_sheet.cell(row_index, 1, key)
        environment_sheet.cell(row_index, 2, value)
        environment_sheet.cell(row_index, 3, "Captured at run time")
        environment_sheet.cell(row_index, 2).font = Font(color=GREEN_TEXT)
    _set_column_widths(environment_sheet, {1: 35, 2: 75, 3: 24, 4: 3})

    # General alignment and print settings ---------------------------------
    for ws in workbook.worksheets:
        ws.freeze_panes = ws.freeze_panes or "A3"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.row != 1:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal or "left",
                        vertical=cell.alignment.vertical or "center",
                        wrap_text=cell.alignment.wrap_text,
                    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


# ---------------------------------------------------------------------------
# Console output and orchestration
# ---------------------------------------------------------------------------
def print_console_table(results: list[BenchmarkResult]) -> None:
    rows = []
    for result in sorted(results, key=lambda item: item.rank):
        rows.append(
            [
                result.rank,
                result.algorithm,
                f"{result.average_latency_ms:.2f}",
                f"{result.peak_memory_mb_per_hash:.2f}",
                f"{result.throughput_hashes_per_second:.3f}",
                f"{result.security_score:.2f}",
                f"{result.latency_score:.2f}",
                f"{result.memory_fit_score:.2f}",
                f"{result.wss:.3f}",
                result.input_length_handling,
            ]
        )

    print("\nPASSWORD HASHING BENCHMARK RESULTS")
    print(
        tabulate(
            rows,
            headers=[
                "Rank",
                "Algorithm",
                "Avg ms/hash",
                "Peak MB/hash",
                "Hashes/sec",
                "Security",
                "Latency",
                "Memory Fit",
                "WSS",
                "Long-input test",
            ],
            tablefmt="grid",
            stralign="left",
            numalign="right",
        )
    )


def run_benchmark(args: argparse.Namespace) -> tuple[list[BenchmarkResult], Path]:
    sample_count = 8 if args.quick else 100
    passwords = generate_dummy_passwords(count=sample_count, seed=args.seed)
    environment = environment_metadata()

    print(f"Generated {len(passwords)} synthetic passwords using seed {args.seed}.")
    print("Running truncation/input-length tests...")
    length_tests = test_input_length_handling()

    raw_latencies: dict[str, list[float]] = {}
    memory_replications: dict[str, list[float]] = {}
    partial_results: list[dict[str, Any]] = []
    representative_password = next(
        (record.password for record in passwords if record.actual_length_chars == 64),
        passwords[-1].password,
    )

    # Fixed order avoids choosing the winner based on execution sequence.
    # The raw workbook records the order and complete measurements.
    for algorithm in ["bcrypt", "Argon2id", "PBKDF2-HMAC-SHA256"]:
        print(f"Benchmarking {algorithm}...")
        check = correctness_check(algorithm, passwords[0].password)
        latencies, total_seconds = benchmark_latencies(algorithm, passwords)
        raw_latencies[algorithm] = latencies

        if args.skip_memory:
            peak_memory_mb = 0.0
            memory_values = [0.0]
        else:
            peak_memory_mb, memory_values = measure_peak_memory_mb(
                algorithm,
                representative_password,
                repetitions=args.memory_repetitions,
                timeout_seconds=args.memory_timeout_seconds,
            )
        memory_replications[algorithm] = memory_values

        avg_latency = statistics.fmean(latencies)
        median_latency = statistics.median(latencies)
        p95_latency = percentile(latencies, 0.95)
        stddev_latency = statistics.stdev(latencies) if len(latencies) > 1 else 0.0
        throughput = len(latencies) / total_seconds if total_seconds > 0 else 0.0
        sec_score = SECURITY_SCORES[algorithm]
        lat_score = latency_score(avg_latency, args.target_latency_ms)
        mem_score = memory_fit_score(peak_memory_mb, args.memory_budget_mb)
        wss = calculate_wss(sec_score, lat_score, mem_score)

        partial_results.append(
            {
                "algorithm": algorithm,
                "configuration": CONFIGURATIONS[algorithm],
                "sample_count": len(latencies),
                "average_latency_ms": avg_latency,
                "median_latency_ms": median_latency,
                "p95_latency_ms": p95_latency,
                "stddev_latency_ms": stddev_latency,
                "min_latency_ms": min(latencies),
                "max_latency_ms": max(latencies),
                "total_time_seconds": total_seconds,
                "throughput_hashes_per_second": throughput,
                "peak_memory_mb_per_hash": peak_memory_mb,
                "security_score": sec_score,
                "latency_score": lat_score,
                "memory_fit_score": mem_score,
                "wss": wss,
                "correctness_check": check,
                "input_length_handling": length_tests[algorithm]["handling"],
                "input_length_details": length_tests[algorithm]["details"],
            }
        )

    ranked = sorted(partial_results, key=lambda item: item["wss"], reverse=True)
    rank_lookup = {item["algorithm"]: index + 1 for index, item in enumerate(ranked)}
    results = [BenchmarkResult(**item, rank=rank_lookup[item["algorithm"]]) for item in partial_results]

    print_console_table(results)
    recommendation = build_recommendation(results, args.target_latency_ms, args.memory_budget_mb)
    print(f"\nRECOMMENDATION\n{recommendation}\n")

    output_path = Path(args.output).expanduser().resolve()
    export_excel_report(
        output_path=output_path,
        results=results,
        raw_latencies=raw_latencies,
        passwords=passwords,
        length_tests=length_tests,
        memory_replications=memory_replications,
        target_latency_ms=args.target_latency_ms,
        memory_budget_mb=args.memory_budget_mb,
        seed=args.seed,
        environment=environment,
        quick_mode=args.quick,
    )
    print(f"Excel audit report created: {output_path}")
    return results, output_path


def parse_arguments(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Benchmark bcrypt, Argon2id, and PBKDF2-HMAC-SHA256 and export an Excel audit report."
    )
    parser.add_argument(
        "--output",
        default="Password_Hashing_Benchmark_Results.xlsx",
        help="Excel output path (default: Password_Hashing_Benchmark_Results.xlsx)",
    )
    parser.add_argument("--seed", type=int, default=2026, help="Synthetic dataset seed")
    parser.add_argument(
        "--target-latency-ms",
        type=float,
        default=350.0,
        help="Policy target used by the latency-fit score (default: 350 ms)",
    )
    parser.add_argument(
        "--memory-budget-mb",
        type=float,
        default=128.0,
        help="Per-hash memory budget used by Memory Fit Score (default: 128 MB)",
    )
    parser.add_argument(
        "--memory-repetitions",
        type=int,
        default=3,
        help="Fresh-process memory replications per algorithm (default: 3)",
    )
    parser.add_argument(
        "--memory-timeout-seconds",
        type=int,
        default=120,
        help="Timeout for each memory worker (default: 120 seconds)",
    )
    parser.add_argument(
        "--skip-memory",
        action="store_true",
        help="Skip RSS memory tests; intended only for troubleshooting",
    )
    parser.add_argument(
        "--quick",
        action="store_true",
        help="Use 8 passwords for a smoke test; not valid for final research results",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_arguments(argv)
    if args.memory_repetitions < 1:
        raise SystemExit("--memory-repetitions must be at least 1")
    if args.target_latency_ms <= 0 or args.memory_budget_mb <= 0:
        raise SystemExit("Latency target and memory budget must be greater than zero")

    try:
        run_benchmark(args)
        return 0
    except KeyboardInterrupt:
        print("Benchmark interrupted by user.", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"Benchmark failed: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    mp.freeze_support()
    raise SystemExit(main())
